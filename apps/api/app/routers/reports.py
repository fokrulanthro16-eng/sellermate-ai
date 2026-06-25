"""
Reports router — CLV, Churn, Revenue Forecast, Health Score, Profit, Tax, PDF/Excel export.
"""
from __future__ import annotations

import io
from dataclasses import asdict

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

from app.ai.commerce import CommerceEngine
from app.core.dependencies import CurrentMerchant, DB
from app.schemas.commerce import (
    BusinessHealthScoreOut,
    ChurnRiskOut,
    CustomerLTVOut,
    HealthComponentOut,
    ProfitReportOut,
    RevenueForecastOut,
    TaxSummaryOut,
)
from app.schemas.common import SuccessResponse

router = APIRouter(tags=["reports"])
_engine = CommerceEngine()


@router.get("/customer-ltv", response_model=SuccessResponse[list[CustomerLTVOut]])
async def customer_ltv(merchant: CurrentMerchant, db: DB):
    data = await _engine.customer_ltv(db, merchant.id)
    return SuccessResponse(data=[CustomerLTVOut(**vars(d)) for d in data])


@router.get("/churn-risk", response_model=SuccessResponse[list[ChurnRiskOut]])
async def churn_risk(merchant: CurrentMerchant, db: DB):
    data = await _engine.churn_risk(db, merchant.id)
    return SuccessResponse(data=[ChurnRiskOut(**vars(d)) for d in data])


@router.get("/revenue-forecast", response_model=SuccessResponse[RevenueForecastOut])
async def revenue_forecast(merchant: CurrentMerchant, db: DB):
    data = await _engine.revenue_forecast(db, merchant.id)
    return SuccessResponse(data=RevenueForecastOut(**vars(data)))


@router.get("/health-score", response_model=SuccessResponse[BusinessHealthScoreOut])
async def health_score(merchant: CurrentMerchant, db: DB):
    data = await _engine.health_score(db, merchant.id)
    return SuccessResponse(data=BusinessHealthScoreOut(
        score=data.score, grade=data.grade,
        components=[HealthComponentOut(**vars(c)) for c in data.components],
        strengths=data.strengths, weaknesses=data.weaknesses,
        explanation_en=data.explanation_en, explanation_bn=data.explanation_bn,
    ))


@router.get("/profit", response_model=SuccessResponse[ProfitReportOut])
async def profit_report(
    merchant: CurrentMerchant,
    db: DB,
    days: int = Query(30, ge=7, le=365),
):
    data = await _engine.profit_report(db, merchant.id, days)
    return SuccessResponse(data=ProfitReportOut(**vars(data)))


@router.get("/tax-summary", response_model=SuccessResponse[TaxSummaryOut])
async def tax_summary(
    merchant: CurrentMerchant,
    db: DB,
    days: int = Query(30, ge=7, le=365),
):
    data = await _engine.tax_summary(db, merchant.id, days)
    return SuccessResponse(data=TaxSummaryOut(**vars(data)))


# ── PDF Export ────────────────────────────────────────────────────────────────

@router.get("/export/pdf")
async def export_pdf(
    merchant: CurrentMerchant,
    db: DB,
    days: int = Query(30, ge=7, le=365),
):
    from fpdf import FPDF

    profit = await _engine.profit_report(db, merchant.id, days)
    tax = await _engine.tax_summary(db, merchant.id, days)
    health = await _engine.health_score(db, merchant.id)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "SellerMate Business Report", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Period: Last {days} days", ln=True, align="C")
    pdf.ln(6)

    def section(title: str):
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, title, ln=True)
        pdf.set_font("Helvetica", "", 10)

    def row(label: str, value: str):
        pdf.cell(90, 6, label, border=0)
        pdf.cell(0, 6, value, ln=True)

    section("Profit Report")
    row("Total Revenue:", f"BDT {profit.total_revenue:,.2f}")
    row("Estimated COGS (60%):", f"BDT {profit.estimated_cogs:,.2f}")
    row("Gross Profit:", f"BDT {profit.gross_profit:,.2f}")
    row("Gross Margin:", f"{profit.gross_margin_pct:.1f}%")
    row("Total Discounts:", f"BDT {profit.total_discounts:,.2f}")
    row("Shipping Cost:", f"BDT {profit.total_shipping_cost:,.2f}")
    row("Net Profit:", f"BDT {profit.net_profit:,.2f}")
    row("Net Margin:", f"{profit.net_margin_pct:.1f}%")
    row("Delivered Orders:", str(profit.delivered_order_count))
    pdf.ln(4)

    section("Tax Summary (Bangladesh)")
    row("VAT Rate:", f"{tax.vat_rate_pct:.0f}%")
    row("Estimated VAT:", f"BDT {tax.estimated_vat:,.2f}")
    row("Estimated Income Tax:", f"BDT {tax.estimated_income_tax:,.2f}")
    row("Total Tax Liability:", f"BDT {tax.total_tax_liability:,.2f}")
    row("Net Tax (after deductions):", f"BDT {tax.net_tax_after_deductions:,.2f}")
    pdf.ln(4)

    section("Business Health Score")
    row("Score:", f"{health.score}/100  (Grade {health.grade})")
    row("Strengths:", ", ".join(health.strengths) or "None")
    row("Weaknesses:", ", ".join(health.weaknesses) or "None")
    pdf.multi_cell(0, 6, health.explanation_en)

    buf = io.BytesIO(bytes(pdf.output()))
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=sellermate_report_{days}d.pdf"},
    )


# ── Excel Export ──────────────────────────────────────────────────────────────

@router.get("/export/excel")
async def export_excel(
    merchant: CurrentMerchant,
    db: DB,
    days: int = Query(30, ge=7, le=365),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    profit = await _engine.profit_report(db, merchant.id, days)
    tax = await _engine.tax_summary(db, merchant.id, days)
    health = await _engine.health_score(db, merchant.id)
    best = await _engine.best_sellers(db, merchant.id, days)
    ltv = await _engine.customer_ltv(db, merchant.id)
    forecast = await _engine.revenue_forecast(db, merchant.id)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Profit ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Profit Report"
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    def write_header(ws, row: int, cols: list[str]):
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")

    write_header(ws, 1, ["Metric", "Value"])
    profit_rows = [
        ("Total Revenue (BDT)", profit.total_revenue),
        ("Estimated COGS (BDT)", profit.estimated_cogs),
        ("Gross Profit (BDT)", profit.gross_profit),
        ("Gross Margin %", profit.gross_margin_pct),
        ("Total Discounts (BDT)", profit.total_discounts),
        ("Shipping Cost (BDT)", profit.total_shipping_cost),
        ("Net Profit (BDT)", profit.net_profit),
        ("Net Margin %", profit.net_margin_pct),
        ("Delivered Orders", profit.delivered_order_count),
        ("Total Orders", profit.total_order_count),
    ]
    for r, (label, val) in enumerate(profit_rows, 2):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # ── Sheet 2: Tax ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Tax Summary")
    write_header(ws2, 1, ["Metric", "Value"])
    tax_rows = [
        ("VAT Rate %", tax.vat_rate_pct),
        ("Estimated VAT (BDT)", tax.estimated_vat),
        ("Gross Profit (BDT)", tax.gross_profit),
        ("Estimated Income Tax (BDT)", tax.estimated_income_tax),
        ("Total Tax Liability (BDT)", tax.total_tax_liability),
        ("Deductible Shipping (BDT)", tax.deductible_shipping),
        ("Deductible Discounts (BDT)", tax.deductible_discounts),
        ("Net Tax After Deductions (BDT)", tax.net_tax_after_deductions),
    ]
    for r, (label, val) in enumerate(tax_rows, 2):
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=val)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 22

    # ── Sheet 3: Health Score ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Health Score")
    write_header(ws3, 1, ["Component", "Score", "Max", "Status"])
    for r, comp in enumerate(health.components, 2):
        ws3.cell(row=r, column=1, value=comp.name)
        ws3.cell(row=r, column=2, value=comp.score)
        ws3.cell(row=r, column=3, value=comp.max_score)
        ws3.cell(row=r, column=4, value=comp.status)
    r = len(health.components) + 3
    ws3.cell(row=r, column=1, value="Overall Score")
    ws3.cell(row=r, column=2, value=health.score)
    ws3.cell(row=r, column=4, value=f"Grade {health.grade}")
    for col in "ABCD":
        ws3.column_dimensions[col].width = 25

    # ── Sheet 4: Best Sellers ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Best Sellers")
    write_header(ws4, 1, ["Product", "Units Sold", "Revenue (BDT)", "Orders", "Avg Price"])
    for r, item in enumerate(best, 2):
        ws4.cell(row=r, column=1, value=item.product_name)
        ws4.cell(row=r, column=2, value=item.total_units)
        ws4.cell(row=r, column=3, value=item.total_revenue)
        ws4.cell(row=r, column=4, value=item.order_count)
        ws4.cell(row=r, column=5, value=item.avg_price)
    for i in range(1, 6):
        ws4.column_dimensions[get_column_letter(i)].width = 22

    # ── Sheet 5: Customer LTV ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("Customer LTV")
    write_header(ws5, 1, ["Name", "Phone", "Orders", "Total Spent (BDT)", "LTV 12m (BDT)", "Segment"])
    for r, c in enumerate(ltv, 2):
        ws5.cell(row=r, column=1, value=c.customer_name)
        ws5.cell(row=r, column=2, value=c.phone)
        ws5.cell(row=r, column=3, value=c.total_orders)
        ws5.cell(row=r, column=4, value=c.total_spent)
        ws5.cell(row=r, column=5, value=c.predicted_ltv_12m)
        ws5.cell(row=r, column=6, value=c.segment)
    for i in range(1, 7):
        ws5.column_dimensions[get_column_letter(i)].width = 22

    # ── Sheet 6: Revenue Forecast ─────────────────────────────────────────────
    ws6 = wb.create_sheet("Revenue Forecast")
    write_header(ws6, 1, ["Date", "Revenue (BDT)"])
    for r, pt in enumerate(forecast.daily_points, 2):
        ws6.cell(row=r, column=1, value=pt["date"])
        ws6.cell(row=r, column=2, value=pt["revenue"])
    r = len(forecast.daily_points) + 3
    ws6.cell(row=r, column=1, value="Predicted Next 30d (BDT)")
    ws6.cell(row=r, column=2, value=forecast.predicted_next_30d)
    ws6.cell(row=r + 1, column=1, value="Trend")
    ws6.cell(row=r + 1, column=2, value=forecast.trend)
    ws6.column_dimensions["A"].width = 20
    ws6.column_dimensions["B"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sellermate_report_{days}d.xlsx"},
    )
